from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any, Iterator

from app.parsing.common import clean_text, heading_rows_from_matrix, infer_value_type, join_non_empty, normalize_decimal
from app.parsing.metadata import detect_period, detect_unit
from app.parsing.models import ParseBundle, ParseIssue, ParsedCell, ParsedDocument, ParsedTable


PARSER_VERSION = "1.0.0"


def _xlsx_merged_maps(worksheet) -> tuple[list[str], dict[tuple[int, int], tuple[int, int]]]:
    ranges: list[str] = []
    anchors: dict[tuple[int, int], tuple[int, int]] = {}
    for merged in worksheet.merged_cells.ranges:
        ranges.append(str(merged))
        anchor = (merged.min_row, merged.min_col)
        for row in range(merged.min_row, merged.max_row + 1):
            for col in range(merged.min_col, merged.max_col + 1):
                anchors[(row, col)] = anchor
    return ranges, anchors


def _matrix_value(matrix: list[list[Any]], row: int, col: int) -> Any:
    if row < 1 or col < 1 or row > len(matrix) or col > len(matrix[row - 1]):
        return None
    return matrix[row - 1][col - 1]


def _headers_for_cell(
    row_values: list[Any],
    col_index: int,
    header_matrix: list[list[Any]],
    header_rows: list[int],
    merged_anchors: dict[tuple[int, int], tuple[int, int]],
) -> tuple[str, str]:
    row_header_candidates = []
    for value in row_values[: max(0, col_index - 1)]:
        if clean_text(value) and normalize_decimal(value) is None:
            row_header_candidates.append(value)
    row_header = join_non_empty(row_header_candidates)
    col_values = []
    for header_row in header_rows:
        anchor_row, anchor_col = merged_anchors.get((header_row, col_index), (header_row, col_index))
        col_values.append(_matrix_value(header_matrix, anchor_row, anchor_col))
    return row_header, join_non_empty(col_values)


def _build_xlsx_cell_factory(
    path: Path,
    document: ParsedDocument,
    sheet_name: str,
    table_id: str,
    header_rows: list[int],
    header_matrix: list[list[Any]],
    merged_anchors: dict[tuple[int, int], tuple[int, int]],
    period: str,
    unit: str,
) -> callable:
    def factory() -> Iterator[ParsedCell]:
        import openpyxl
        from openpyxl.utils import get_column_letter

        formula_book = openpyxl.load_workbook(path, read_only=True, data_only=False)
        value_book = openpyxl.load_workbook(path, read_only=True, data_only=True)
        formula_sheet = formula_book[sheet_name]
        value_sheet = value_book[sheet_name]
        try:
            value_rows = value_sheet.iter_rows()
            for row_index, formula_row in enumerate(formula_sheet.iter_rows(), start=1):
                cached_row = next(value_rows, ())
                formula_values = [cell.value for cell in formula_row]
                cached_values = [cell.value for cell in cached_row]
                width = max(len(formula_values), len(cached_values))
                if width == 0:
                    continue
                formula_values.extend([None] * (width - len(formula_values)))
                cached_values.extend([None] * (width - len(cached_values)))
                for col_index, raw in enumerate(formula_values, start=1):
                    cached = cached_values[col_index - 1]
                    if raw is None and cached is None:
                        continue
                    formula = clean_text(raw) if isinstance(raw, str) and raw.startswith("=") else ""
                    display = clean_text(cached if cached is not None else raw)
                    raw_text = clean_text(raw)
                    row_header, col_header = _headers_for_cell(
                        formula_values, col_index, header_matrix, header_rows, merged_anchors
                    )
                    is_header = row_index in set(header_rows)
                    anchor = merged_anchors.get((row_index, col_index))
                    anchor_ref = ""
                    if anchor:
                        anchor_ref = f"{get_column_letter(anchor[1])}{anchor[0]}"
                    cell_ref = f"{get_column_letter(col_index)}{row_index}"
                    cell_period = detect_period(col_header, row_header, period)
                    cell_unit = detect_unit(col_header, row_header, unit)
                    metric_name = row_header or col_header
                    yield ParsedCell(
                        cell_id=f"{table_id}_r{row_index}_c{col_index}",
                        table_id=table_id,
                        doc_id=document.doc_id,
                        row_index=row_index,
                        col_index=col_index,
                        cell_ref=cell_ref,
                        raw_value=raw_text,
                        display_value=display,
                        normalized_value=normalize_decimal(cached if cached is not None else raw),
                        value_type=infer_value_type(cached if cached is not None else raw, formula),
                        formula=formula,
                        metric_name=metric_name,
                        period=cell_period,
                        unit=cell_unit,
                        row_header=row_header,
                        col_header=col_header,
                        is_header=is_header,
                        is_merged=anchor is not None,
                        merged_anchor_ref=anchor_ref,
                        source_locator={"sheet_name": sheet_name, "cell_ref": cell_ref},
                    )
        finally:
            formula_book.close()
            value_book.close()

    return factory


def parse_xlsx(path: Path, document: ParsedDocument) -> ParseBundle:
    import openpyxl
    from openpyxl.utils import get_column_letter

    workbook = openpyxl.load_workbook(path, read_only=False, data_only=False)
    tables: list[ParsedTable] = []
    issues: list[ParseIssue] = []
    try:
        document.sheet_count = len(workbook.sheetnames)
        document.parser_name = "openpyxl"
        document.parser_version = PARSER_VERSION
        for sheet_index, worksheet in enumerate(workbook.worksheets, start=1):
            merged_ranges, merged_anchors = _xlsx_merged_maps(worksheet)
            scan_rows = min(30, worksheet.max_row)
            scan_cols = min(100, worksheet.max_column)
            header_matrix = [
                [worksheet.cell(row, col).value for col in range(1, scan_cols + 1)]
                for row in range(1, scan_rows + 1)
            ]
            header_rows = heading_rows_from_matrix(header_matrix)
            top_text = join_non_empty(value for row in header_matrix[:10] for value in row)
            table_name = next(
                (clean_text(value) for row in header_matrix[:5] for value in row if clean_text(value)),
                worksheet.title,
            )
            unit = detect_unit(top_text)
            period = detect_period(document.title, document.file_name, worksheet.title, top_text)
            table_id = f"{document.doc_id}_sheet_{sheet_index:03d}"
            range_ref = f"A1:{get_column_letter(worksheet.max_column)}{worksheet.max_row}"
            tables.append(
                ParsedTable(
                    table_id=table_id,
                    doc_id=document.doc_id,
                    sequence_no=sheet_index,
                    source_kind="xlsx",
                    table_index=sheet_index,
                    table_name=table_name,
                    sheet_name=worksheet.title,
                    range_ref=range_ref,
                    unit=unit,
                    period=period,
                    header_rows=header_rows,
                    row_count=worksheet.max_row,
                    column_count=worksheet.max_column,
                    merged_ranges=merged_ranges,
                    source_locator={"sheet_name": worksheet.title, "range_ref": range_ref},
                    cells_factory=_build_xlsx_cell_factory(
                        path,
                        document,
                        worksheet.title,
                        table_id,
                        header_rows,
                        header_matrix,
                        merged_anchors,
                        period,
                        unit,
                    ),
                )
            )
    finally:
        workbook.close()
    return ParseBundle(document=document, tables=tables, issues=issues)


def _xlrd_value(book, cell) -> Any:
    import xlrd

    if cell.ctype == xlrd.XL_CELL_DATE:
        return datetime(*xlrd.xldate_as_tuple(cell.value, book.datemode))
    if cell.ctype == xlrd.XL_CELL_BOOLEAN:
        return bool(cell.value)
    return cell.value


def _build_xls_cell_factory(
    path: Path,
    document: ParsedDocument,
    sheet_name: str,
    table_id: str,
    header_rows: list[int],
    header_matrix: list[list[Any]],
    merged_anchors: dict[tuple[int, int], tuple[int, int]],
    period: str,
    unit: str,
) -> callable:
    def factory() -> Iterator[ParsedCell]:
        import xlrd

        book = xlrd.open_workbook(str(path), on_demand=True, formatting_info=False)
        sheet = book.sheet_by_name(sheet_name)
        try:
            for row_index in range(1, sheet.nrows + 1):
                row_values = [_xlrd_value(book, sheet.cell(row_index - 1, col)) for col in range(sheet.ncols)]
                for col_index, raw in enumerate(row_values, start=1):
                    if clean_text(raw) == "":
                        continue
                    row_header, col_header = _headers_for_cell(
                        row_values, col_index, header_matrix, header_rows, merged_anchors
                    )
                    anchor = merged_anchors.get((row_index, col_index))
                    anchor_ref = xlrd.formula.cellname(anchor[0] - 1, anchor[1] - 1) if anchor else ""
                    cell_ref = xlrd.formula.cellname(row_index - 1, col_index - 1)
                    yield ParsedCell(
                        cell_id=f"{table_id}_r{row_index}_c{col_index}",
                        table_id=table_id,
                        doc_id=document.doc_id,
                        row_index=row_index,
                        col_index=col_index,
                        cell_ref=cell_ref,
                        raw_value=clean_text(raw),
                        display_value=clean_text(raw),
                        normalized_value=normalize_decimal(raw),
                        value_type=infer_value_type(raw),
                        metric_name=row_header or col_header,
                        period=detect_period(col_header, row_header, period),
                        unit=detect_unit(col_header, row_header, unit),
                        row_header=row_header,
                        col_header=col_header,
                        is_header=row_index in set(header_rows),
                        is_merged=anchor is not None,
                        merged_anchor_ref=anchor_ref,
                        source_locator={"sheet_name": sheet_name, "cell_ref": cell_ref},
                    )
        finally:
            book.release_resources()

    return factory


def parse_xls(path: Path, document: ParsedDocument) -> ParseBundle:
    import xlrd

    workbook = xlrd.open_workbook(str(path), on_demand=True, formatting_info=True)
    tables: list[ParsedTable] = []
    issues: list[ParseIssue] = []
    document.sheet_count = workbook.nsheets
    document.parser_name = "xlrd"
    document.parser_version = PARSER_VERSION
    try:
        for sheet_index, sheet_name in enumerate(workbook.sheet_names(), start=1):
            sheet = workbook.sheet_by_name(sheet_name)
            scan_rows = min(30, sheet.nrows)
            scan_cols = min(100, sheet.ncols)
            header_matrix = [
                [_xlrd_value(workbook, sheet.cell(row, col)) for col in range(scan_cols)]
                for row in range(scan_rows)
            ]
            header_rows = heading_rows_from_matrix(header_matrix)
            merged_ranges: list[str] = []
            merged_anchors: dict[tuple[int, int], tuple[int, int]] = {}
            for row_low, row_high, col_low, col_high in sheet.merged_cells:
                anchor = (row_low + 1, col_low + 1)
                merged_ranges.append(
                    f"{xlrd.formula.cellname(row_low, col_low)}:{xlrd.formula.cellname(row_high - 1, col_high - 1)}"
                )
                for row in range(row_low + 1, row_high + 1):
                    for col in range(col_low + 1, col_high + 1):
                        merged_anchors[(row, col)] = anchor
            top_text = join_non_empty(value for row in header_matrix[:10] for value in row)
            table_name = next(
                (clean_text(value) for row in header_matrix[:5] for value in row if clean_text(value)),
                sheet_name,
            )
            unit = detect_unit(top_text)
            period = detect_period(document.title, document.file_name, sheet_name, top_text)
            table_id = f"{document.doc_id}_sheet_{sheet_index:03d}"
            range_ref = (
                f"A1:{xlrd.formula.cellname(max(0, sheet.nrows - 1), max(0, sheet.ncols - 1))}"
                if sheet.nrows and sheet.ncols
                else ""
            )
            tables.append(
                ParsedTable(
                    table_id=table_id,
                    doc_id=document.doc_id,
                    sequence_no=sheet_index,
                    source_kind="xls",
                    table_index=sheet_index,
                    table_name=table_name,
                    sheet_name=sheet_name,
                    range_ref=range_ref,
                    unit=unit,
                    period=period,
                    header_rows=header_rows,
                    row_count=sheet.nrows,
                    column_count=sheet.ncols,
                    merged_ranges=merged_ranges,
                    source_locator={"sheet_name": sheet_name, "range_ref": range_ref},
                    cells_factory=_build_xls_cell_factory(
                        path,
                        document,
                        sheet_name,
                        table_id,
                        header_rows,
                        header_matrix,
                        merged_anchors,
                        period,
                        unit,
                    ),
                )
            )
    finally:
        workbook.release_resources()
    return ParseBundle(document=document, tables=tables, issues=issues)


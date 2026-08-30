# -*- coding: utf-8 -*-
"""Batch QA evaluation through the frontend's HTTP endpoint."""
from __future__ import annotations
import argparse, csv, json, re, time
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from urllib import error, request
from openpyxl import load_workbook
try:
    from scripts.qa_eval_common import (
        DEFAULT_LABEL_CORRECTIONS_PATH,
        actual_behavior,
        apply_label_corrections,
        build_behavior_metrics,
        build_retrieval_metrics,
        expected_behavior,
        load_label_corrections,
        retrieval_coverage,
    )
except ModuleNotFoundError:  # direct execution: python scripts/evaluate_qa_http.py
    from qa_eval_common import (
        DEFAULT_LABEL_CORRECTIONS_PATH,
        actual_behavior,
        apply_label_corrections,
        build_behavior_metrics,
        build_retrieval_metrics,
        expected_behavior,
        load_label_corrections,
        retrieval_coverage,
    )

ROOT = Path(__file__).resolve().parents[1]
KEYS = ("A", "B", "C", "D")

def cell(row, key):
    v = row.get(key)
    return "" if v is None else str(v).strip()

def build_query(row):
    q = cell(row, "question")
    opts = [f"{k}. {cell(row, 'option_' + k.lower())}" for k in KEYS if cell(row, 'option_' + k.lower())]
    if opts and not re.search(r"(?:^|\s)[ABCD][.、:：]", q):
        q += "\n" + "\n".join(opts)
    return q

def original_row(row):
    restored = dict(row)
    for key in ("question", "option_a", "option_b", "option_c", "option_d", "answer", "answer_text", "evidence"):
        original_key = f"original_{key}"
        if original_key in row:
            restored[key] = row[original_key]
    return restored

def letters(v):
    return list(dict.fromkeys(re.findall(r"[ABCD]", str(v or "").upper())))

def load_rows(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    headers = [str(v).strip() if v is not None else "" for v in next(it)]
    rows = []
    for raw in it:
        row = {headers[i]: raw[i] if i < len(raw) else None for i in range(len(headers))}
        if cell(row, "question"):
            rows.append(row)
    return rows

def normalize_question_id(value):
    text = str(value or "").strip().upper()
    match = re.fullmatch(r"Q?0*(\d+)", text)
    return f"Q{int(match.group(1)):03d}" if match else text

def parse_id_list(value):
    return {
        normalize_question_id(item)
        for item in re.split(r"[,，\s]+", value or "")
        if item.strip()
    }

def parse_ranges(values):
    selected = set()
    for value in values or []:
        for spec in re.split(r"[,，\s]+", value):
            if not spec:
                continue
            match = re.fullmatch(r"Q?0*(\d+)\s*[-~～至]\s*Q?0*(\d+)", spec, re.I)
            if not match:
                raise ValueError(f"无效题号范围: {spec}，示例: Q010-Q020")
            start, end = map(int, match.groups())
            if start > end:
                start, end = end, start
            selected.update(f"Q{number:03d}" for number in range(start, end + 1))
    return selected

def load_failed_ids(path):
    """Load wrong/refused/error question IDs from a previous evaluation result."""
    if path is None:
        return set()
    if path.suffix.lower() == ".jsonl":
        records = [json.loads(line) for line in path.read_text(encoding="utf-8").splitlines() if line.strip()]
    else:
        payload = json.loads(path.read_text(encoding="utf-8"))
        if isinstance(payload, dict) and any(key in payload for key in ("wrong_ids", "refused_ids", "failed_ids")):
            ids = []
            for key in ("wrong_ids", "refused_ids", "failed_ids"):
                ids.extend(payload.get(key) or [])
            return {normalize_question_id(item) for item in ids}
        records = payload.get("results", []) if isinstance(payload, dict) else payload

    failed = set()
    for item in records:
        if not isinstance(item, dict):
            continue
        status = str(item.get("status") or "").lower()
        has_problem = (
            item.get("accurate") is False
            or item.get("behavior_correct") is False
            or bool(item.get("refused"))
            or status in {"request_error", "http_error", "error", "unknown"}
        )
        if has_problem and item.get("id"):
            failed.add(normalize_question_id(item["id"]))
    return failed

def select_rows(rows, *, ids=None, ranges=None, failed_ids=None, limit=None, selection_requested=False):
    requested = parse_id_list(ids) | parse_ranges(ranges) | set(failed_ids or ())
    selected = rows
    if requested or selection_requested:
        selected = [row for row in rows if normalize_question_id(cell(row, "id")) in requested]
        found = {normalize_question_id(cell(row, "id")) for row in selected}
        missing = sorted(requested - found)
        if missing:
            print(f"警告: QA 文件中未找到题号: {', '.join(missing)}", flush=True)
    return selected[:limit] if limit and limit > 0 else selected

def post_json(url, payload, timeout):
    req = request.Request(url, data=json.dumps(payload, ensure_ascii=False).encode("utf-8"), headers={"Content-Type": "application/json", "Accept": "application/json"}, method="POST")
    try:
        with request.urlopen(req, timeout=timeout) as r:
            return json.loads(r.read().decode("utf-8", errors="replace"))
    except error.HTTPError as e:
        return {"status": "http_error", "error": f"HTTP {e.code}: {e.read().decode('utf-8', errors='replace')[:500]}"}
    except Exception as e:
        return {"status": "request_error", "error": f"{type(e).__name__}: {e}"}

def predicted(result):
    text = str(result.get("answer") or "")
    answer_match = re.search(
        r"^\s*(?:(?:答案|选项|正确选项|选择)\s*[:：]?\s*)?"
        r"(?:\*\*)?\s*([ABCD])(?=\s|[.。、:：\[\]*]|$)",
        text,
        re.I,
    )
    if answer_match:
        return [answer_match.group(1).upper()]

    ver = result.get("verification") if isinstance(result, dict) else None
    if isinstance(ver, dict):
        te = ver.get("table_execution")
        if isinstance(te, dict) and te.get("matched_option"):
            return letters(te["matched_option"])
        ov = ver.get("option_verification")
        if isinstance(ov, dict) and ov.get("selected_options"):
            return letters(ov["selected_options"])
    m = re.findall(r"(?:答案|选项|正确选项|选择)\s*[:：]?\s*([ABCD])(?:\b|[.。、])", text, re.I)
    if m:
        return list(dict.fromkeys(x.upper() for x in m))
    m = re.fullmatch(r"\s*([ABCD])\s*[.。、]?\s*", text, re.I)
    return [m.group(1).upper()] if m else []

def reason(result):
    for k in ("refusal_reason", "error_code"):
        if result.get(k):
            return str(result[k])
    v = result.get("verification")
    if isinstance(v, dict):
        if v.get("error_code"):
            return str(v["error_code"])
        if v.get("issues"):
            return "; ".join(map(str, v["issues"]))
    return str(result.get("error") or "")

def _evaluate_one(row, index, endpoint, timeout):
    q = build_query(row)
    original = original_row(row)
    original_query = build_query(original)
    started = time.perf_counter()
    resp = post_json(endpoint, {"question": q, "top_k": 5}, timeout)
    ms = int((time.perf_counter() - started) * 1000)
    gold = letters(row.get("answer")) or letters(row.get("answer_text"))
    original_gold = letters(original.get("answer")) or letters(original.get("answer_text"))
    pred = predicted(resp)
    status = str(resp.get("status") or "unknown")
    expected = expected_behavior(row)
    actual = actual_behavior(status)
    refused = status not in {"answered", "success", "degraded"}
    accurate = bool(gold) and pred == gold and not refused
    item = {
        "id": cell(row, "id") or f"row_{index:03d}",
        "qa_type": cell(row, "qa_type"),
        "question": cell(row, "question"),
        "original_question": cell(original, "question"),
        "original_submitted_question": original_query,
        "submitted_question": q,
        "expected_behavior": expected,
        "actual_behavior": actual,
        "behavior_correct": expected == actual,
        "retrieval_coverage": retrieval_coverage(resp),
        "original_options": {key: cell(original, f"option_{key.lower()}") for key in KEYS},
        "options": {key: cell(row, f"option_{key.lower()}") for key in KEYS},
        "gold_answer_original": original_gold,
        "gold_answer_effective": gold,
        "gold_answer": gold,
        "gold_answer_text_original": cell(original, "answer_text"),
        "gold_answer_text_effective": cell(row, "answer_text"),
        "gold_evidence_original": cell(original, "evidence"),
        "gold_evidence_effective": cell(row, "evidence"),
        "correction_applied": bool(row.get("label_correction_applied")),
        "correction_version": str(row.get("label_correction_version") or ""),
        "correction_reason": str(row.get("label_correction_reason") or ""),
        "correction_fields": list(row.get("label_correction_fields") or []),
        "correction_source_cells": list(row.get("label_correction_source_cells") or []),
        "predicted_answer": pred,
        "status": status,
        "refused": refused,
        "accurate": accurate,
        "refusal_reason": reason(resp),
        "source_title": cell(row, "source_title"),
        "system_answer": resp.get("answer", ""),
        "elapsed_ms": ms,
        "response": resp,
    }
    return index, item

def evaluate(rows, endpoint, limit, timeout, workers=1):
    selected = rows[:limit] if limit and limit > 0 else rows
    completed = []
    worker_count = max(1, int(workers))

    if worker_count == 1:
        iterator = (_evaluate_one(row, index, endpoint, timeout) for index, row in enumerate(selected, 1))
        for done, result in enumerate(iterator, 1):
            index, item = result
            completed.append((index, item))
            print(f"[{done}/{len(selected)}] {item['id']} status={item['status']} pred={item['predicted_answer'] or '-'} gold={item['gold_answer'] or '-'} accurate={item['accurate']} {item['elapsed_ms']}ms", flush=True)
    else:
        with ThreadPoolExecutor(max_workers=worker_count) as executor:
            futures = [
                executor.submit(_evaluate_one, row, index, endpoint, timeout)
                for index, row in enumerate(selected, 1)
            ]
            for done, future in enumerate(as_completed(futures), 1):
                index, item = future.result()
                completed.append((index, item))
                print(f"[{done}/{len(selected)}] {item['id']} status={item['status']} pred={item['predicted_answer'] or '-'} gold={item['gold_answer'] or '-'} accurate={item['accurate']} {item['elapsed_ms']}ms", flush=True)

    completed.sort(key=lambda pair: pair[0])
    return [item for _, item in completed]

def report(results, outdir, endpoint):
    outdir.mkdir(parents=True, exist_ok=True)
    total = len(results); refused = sum(x["refused"] for x in results); answered = total - refused; correct = sum(x["accurate"] for x in results)
    wrong = [x for x in results if not x["refused"] and not x["accurate"]]; refusals = [x for x in results if x["refused"]]
    by_type = {}
    for x in results:
        b = by_type.setdefault(x["qa_type"] or "unclassified", {"total": 0, "answered": 0, "refused": 0, "correct": 0, "wrong": 0})
        b["total"] += 1; b["answered"] += int(not x["refused"]); b["refused"] += int(x["refused"]); b["correct"] += int(x["accurate"]); b["wrong"] += int(not x["refused"] and not x["accurate"])
    corrected = [x for x in results if x.get("correction_applied")]
    summary = {"total": total, "question_ids": [x["id"] for x in results], "answered": answered, "refused": refused, "correct": correct, "wrong_answered": len(wrong), "accuracy_over_all": correct / total if total else 0, "accuracy_over_answered": correct / answered if answered else 0, "label_corrections": {"applied_count": len(corrected), "question_ids": [x["id"] for x in corrected], "versions": sorted({x["correction_version"] for x in corrected if x.get("correction_version")})}, "retrieval_metrics": build_retrieval_metrics(results), "behavior_metrics": build_behavior_metrics(results), "refused_ids": [x["id"] for x in refusals], "wrong_ids": [x["id"] for x in wrong], "by_qa_type": by_type}
    (outdir / "qa_summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    with (outdir / "qa_results.jsonl").open("w", encoding="utf-8") as f:
        for x in results: f.write(json.dumps(x, ensure_ascii=False) + "\n")
    cols = ["id", "qa_type", "original_question", "original_submitted_question", "submitted_question", "expected_behavior", "actual_behavior", "behavior_correct", "gold_answer_original", "gold_answer_effective", "gold_answer", "correction_applied", "correction_version", "correction_reason", "predicted_answer", "status", "refused", "accurate", "refusal_reason", "source_title", "system_answer", "elapsed_ms"]
    with (outdir / "qa_results.csv").open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=cols, extrasaction="ignore"); w.writeheader(); w.writerows(results)
    behavior = summary["behavior_metrics"]
    retrieval = summary["retrieval_metrics"]
    md = ["# QA HTTP Evaluation Summary", "", f"- Endpoint: `POST {endpoint}` (same as frontend)", "- Sent only question and options; gold answer/evidence were not sent.", "- DeepSeek remained enabled according to `.env`; local mode was not used.", f"- Scope: {total} selected questions; accuracy applies only to this run.", f"- Label corrections applied: {len(corrected)}" + (f" ({', '.join(x['id'] for x in corrected)})" if corrected else ""), f"- Total: {total}", f"- Answered: {answered}", f"- Refused/request failed: {refused}", f"- Correct: {correct}", f"- Wrong among answered: {len(wrong)}", f"- Overall accuracy: {correct / total:.2%}" if total else "- Overall accuracy: 0.00%", f"- Answered accuracy: {correct / answered:.2%}" if answered else "- Answered accuracy: 0.00%", f"- Average target coverage: {retrieval['average_target_coverage']:.2%}" if retrieval["average_target_coverage"] is not None else "- Average target coverage: n/a", f"- Full target coverage rate: {retrieval['full_coverage_rate']:.2%}" if retrieval["full_coverage_rate"] is not None else "- Full target coverage rate: n/a", f"- Behavior accuracy: {behavior['behavior_accuracy']:.2%}" if behavior["behavior_accuracy"] is not None else "- Behavior accuracy: n/a", f"- False refusal rate: {behavior['false_refusal_rate']:.2%}" if behavior["false_refusal_rate"] is not None else "- False refusal rate: n/a", f"- Refusal precision/recall: {behavior['refusal_precision']:.2%} / {behavior['refusal_recall']:.2%}" if behavior["refusal_precision"] is not None and behavior["refusal_recall"] is not None else "- Refusal precision/recall: n/a", f"- Clarification precision/recall: {behavior['clarification_precision']:.2%} / {behavior['clarification_recall']:.2%}" if behavior["clarification_precision"] is not None and behavior["clarification_recall"] is not None else "- Clarification precision/recall: n/a", "", "## By type", "", "|Type|Total|Answered|Refused|Correct|Wrong|", "|---|---:|---:|---:|---:|---:|"]
    for typ, b in by_type.items():
        md.append(f"|{typ}|{b['total']}|{b['answered']}|{b['refused']}|{b['correct']}|{b['wrong']}|")
    md += ["", "## Refused/request failed", "", "|ID|Status|Reason|", "|---|---|---|"]
    md += [f"|{x['id']}|{x['status']}|{x['refusal_reason'] or 'unspecified'}|" for x in refusals] or ["|none| | |"]
    md += ["", "## Wrong answered", "", "|ID|Gold|Predicted|Status|", "|---|---|---|---|"]
    md += [f"|{x['id']}|{''.join(x['gold_answer'])}|{''.join(x['predicted_answer']) or 'unparsed'}|{x['status']}|" for x in wrong] or ["|none| | | |"]
    (outdir / "QA_HTTP_summary.md").write_text("\n".join(md) + "\n", encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--qa", type=Path, default=Path(r"D:\金融科技\QA数据.xlsx"))
    ap.add_argument("--endpoint", default="http://127.0.0.1:8000/api/v1/ask")
    ap.add_argument("--output-dir", type=Path, default=None, help="默认完整测试写 qa_eval_http_full，筛选测试写 qa_eval_http_subset")
    ap.add_argument("--limit", type=int, default=None, help="筛选后最多测试前 N 题")
    ap.add_argument("--timeout", type=float, default=180)
    ap.add_argument("--workers", type=int, default=3, help="并发请求数；设为 1 可恢复串行")
    ap.add_argument("--ids", default="", help="指定题号，如 Q010,Q015,Q020")
    ap.add_argument("--range", dest="ranges", action="append", default=[], help="题号闭区间，如 Q010-Q020；可重复")
    ap.add_argument("--failed-from", type=Path, default=None, help="只重测旧 qa_results.jsonl/qa_summary.json 中的问题题")
    ap.add_argument("--corrections", type=Path, default=DEFAULT_LABEL_CORRECTIONS_PATH, help="QA 标签修正覆盖文件")
    ap.add_argument("--no-label-corrections", action="store_true", help="不应用修正，用于复现原始 QA 基线")
    args = ap.parse_args()

    if args.workers < 1:
        ap.error("--workers 必须大于等于 1")

    selection_requested = bool(args.ids.strip() or args.ranges or args.failed_from)
    try:
        failed_ids = load_failed_ids(args.failed_from)
        rows = load_rows(args.qa)
        if not args.no_label_corrections:
            rows = apply_label_corrections(rows, load_label_corrections(args.corrections))
        selected = select_rows(
            rows,
            ids=args.ids,
            ranges=args.ranges,
            failed_ids=failed_ids,
            limit=args.limit,
            selection_requested=selection_requested,
        )
    except (OSError, ValueError) as exc:
        ap.error(str(exc))

    if not selected:
        ap.error("筛选后没有可测试题目")
    output_dir = args.output_dir or ROOT / "reports" / (
        "qa_eval_http_subset" if selection_requested else "qa_eval_http_full"
    )
    print(f"将测试 {len(selected)} 题，并发数={max(1, args.workers)}", flush=True)
    report(
        evaluate(selected, args.endpoint, None, args.timeout, args.workers),
        output_dir,
        args.endpoint,
    )

if __name__ == "__main__":
    main()

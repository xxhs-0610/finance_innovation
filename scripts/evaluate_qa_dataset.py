"""Evaluate the RAG system against the curated QA Excel workbook.

The workbook supplies the reference answer/options/evidence.  Only the
question plus its options are sent to the RAG system; the evidence column is
used as a gold reference in the report and is never leaked into the query.

Example (from the project root)::

    python scripts/evaluate_qa_dataset.py --qa "D:\\金融科技\\QA数据.xlsx" \
        --local --limit 20

The script writes ``qa_results.jsonl``, ``qa_results.csv`` and
``qa_summary.json`` under ``reports/qa_eval`` by default.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
import time
from pathlib import Path
from typing import Any

try:
    from scripts.qa_eval_common import actual_behavior, build_behavior_metrics, build_retrieval_metrics, expected_behavior, retrieval_coverage
except ModuleNotFoundError:  # direct execution: python scripts/evaluate_qa_dataset.py
    from qa_eval_common import actual_behavior, build_behavior_metrics, build_retrieval_metrics, expected_behavior, retrieval_coverage

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover - clear message for setup issues
    raise SystemExit("缺少 openpyxl，请在 ssgs 环境安装项目 requirements.txt") from exc


PROJECT_ROOT = Path(__file__).resolve().parents[1]
BACKEND_ROOT = PROJECT_ROOT / "backend"
for path in (PROJECT_ROOT, BACKEND_ROOT):
    if str(path) not in sys.path:
        sys.path.insert(0, str(path))


OPTION_KEYS = ("A", "B", "C", "D")
STATUS_ANSWERED = {"answered", "success"}


def _default_qa_path() -> Path:
    candidates = (
        PROJECT_ROOT / "data" / "raw" / "QA数据.xlsx",
        PROJECT_ROOT / "QA数据.xlsx",
        PROJECT_ROOT.parent.parent / "QA数据.xlsx",
    )
    return next((p for p in candidates if p.exists()), candidates[0])


def _cell(row: dict[str, Any], name: str) -> str:
    value = row.get(name)
    return "" if value is None else str(value).strip()


def _build_query(row: dict[str, Any]) -> str:
    """Add options to the original question without adding gold evidence."""
    question = _cell(row, "question")
    options = []
    for key in OPTION_KEYS:
        value = _cell(row, f"option_{key.lower()}")
        if value:
            options.append(f"{key}. {value}")
    if not options:
        return question
    # Avoid duplicating options when a source workbook already embeds them.
    if any(re.search(rf"(?:^|\s){key}[.:、：]", question) for key in OPTION_KEYS):
        return question
    return question + "\n" + "\n".join(options)


def _normalise_letters(value: Any) -> list[str]:
    text = "" if value is None else str(value).upper()
    found = re.findall(r"[ABCD]", text)
    # Preserve order while removing duplicates.
    return list(dict.fromkeys(found))


def _extract_system_letters(result: dict[str, Any]) -> list[str]:
    # Deterministic table execution is authoritative for Excel questions.
    # Prefer its matched option before any free-form answer parsing.
    verification = result.get("verification")
    if isinstance(verification, dict):
        table_execution = verification.get("table_execution")
        if isinstance(table_execution, dict) and table_execution.get("matched_option"):
            return _normalise_letters(table_execution["matched_option"])
    if isinstance(verification, dict):
        option_verification = verification.get("option_verification")
        if isinstance(option_verification, dict):
            selected = option_verification.get("selected_options")
            letters = _normalise_letters(selected)
            if letters:
                return letters
    # Only accept explicit answer labels.  Do not scan the entire answer for
    # A/B/C/D: evidence often contains cell references such as ``B36:G36``.
    answer_text = str(result.get("answer") or "")
    explicit = re.findall(r"(?:答案|选项|正确选项|选择)\s*[:：]?\s*([ABCD])(?:\b|[.。、])", answer_text, flags=re.I)
    if explicit:
        return list(dict.fromkeys(x.upper() for x in explicit))
    # A bare one-letter answer is also safe, but only when the complete
    # response is that letter (possibly surrounded by Markdown).
    bare = re.fullmatch(r"\s*[（(]?\s*([ABCD])\s*[）)]?\s*[。.]?\s*", answer_text, flags=re.I)
    return [bare.group(1).upper()] if bare else []


def _normalise_number(text: Any) -> str:
    """Normalize a numeric option for exact textual comparison."""
    value = "" if text is None else str(text).strip()
    value = value.replace(",", "").replace("，", "")
    # Keep decimals/signs; strip trailing zeroes only for decimal values.
    try:
        number = float(value)
    except (TypeError, ValueError):
        return value
    return f"{number:.12g}"


def _infer_answer_from_option_values(result: dict[str, Any], row: dict[str, Any]) -> list[str]:
    """Infer option letters when the generated answer contains only a value."""
    answer_text = str(result.get("answer") or "")
    if not answer_text:
        return []
    matches: list[str] = []
    for key in OPTION_KEYS:
        option_value = _cell(row, f"option_{key.lower()}")
        if not option_value:
            continue
        # Match the normalized numeric token, allowing units or Markdown around it.
        normalized = _normalise_number(option_value)
        if normalized and normalized in answer_text.replace(",", "").replace("，", ""):
            matches.append(key)
    # A generated table explanation may contain every candidate value.  Such
    # multi-match output is ambiguous and must never be treated as a selected
    # answer; structured table_execution/option_verification is required.
    return matches if len(matches) == 1 else []


def _refusal_reason(result: dict[str, Any]) -> str:
    for key in ("refusal_reason", "error_code"):
        if result.get(key):
            return str(result[key])
    verification = result.get("verification")
    if isinstance(verification, dict):
        if verification.get("error_code"):
            return str(verification["error_code"])
        issues = verification.get("issues")
        if issues:
            return "; ".join(map(str, issues))
    diagnostics = result.get("diagnostics")
    if isinstance(diagnostics, dict):
        router = diagnostics.get("router")
        if isinstance(router, dict) and router.get("reason"):
            return str(router["reason"])
    return ""


def _option_features(result: dict[str, Any]) -> dict[str, Any]:
    verification = result.get("verification")
    if not isinstance(verification, dict):
        return {}
    option_verification = verification.get("option_verification")
    if not isinstance(option_verification, dict):
        return {}
    features: dict[str, Any] = {}
    for item in option_verification.get("options_verification", []) or []:
        if not isinstance(item, dict):
            continue
        option = str(item.get("option") or "")
        if option:
            features[option] = {
                key: item.get(key)
                for key in ("R_i", "E_i", "M_i", "Delta_i", "N_i", "C_i", "verdict", "similarity")
                if key in item
            }
    return features


def _gold_answer(row: dict[str, Any]) -> list[str]:
    letters = _normalise_letters(row.get("answer"))
    if letters:
        return letters
    return _normalise_letters(row.get("answer_text"))


def load_rows(path: Path, sheet: str | None = None) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    selected_sheet = sheet or workbook.sheetnames[0]
    if selected_sheet not in workbook.sheetnames:
        raise ValueError(f"找不到工作表 {selected_sheet!r}，可选：{workbook.sheetnames}")
    ws = workbook[selected_sheet]
    values = ws.iter_rows(values_only=True)
    headers = [str(v).strip() if v is not None else "" for v in next(values)]
    rows: list[dict[str, Any]] = []
    for raw in values:
        row = {headers[i]: raw[i] if i < len(raw) else None for i in range(len(headers))}
        if _cell(row, "question"):
            rows.append(row)
    return rows


def evaluate(rows: list[dict[str, Any]], *, local: bool, limit: int | None, delay: float) -> list[dict[str, Any]]:
    from app.services.rag_service import RAGService

    service = RAGService()
    selected_rows = rows[:limit] if limit and limit > 0 else rows
    results: list[dict[str, Any]] = []
    for index, row in enumerate(selected_rows, start=1):
        query = _build_query(row)
        started = time.perf_counter()
        try:
            kwargs: dict[str, Any] = {}
            if local:
                kwargs["deepseek_enabled_fn"] = lambda: False
            response = service.ask(query, **kwargs)
            if not isinstance(response, dict):
                response = {"status": "error", "answer": str(response)}
            error = ""
        except Exception as exc:  # keep the batch running and report failures
            response = {"status": "error", "answer": "", "error": repr(exc)}
            error = repr(exc)
        elapsed_ms = int((time.perf_counter() - started) * 1000)

        gold = _gold_answer(row)
        # Structured execution/verification is authoritative.  Numeric value
        # inference is only a last resort and accepts exactly one match.
        predicted = _extract_system_letters(response)
        if not predicted:
            predicted = _infer_answer_from_option_values(response, row)
        status = str(response.get("status") or "unknown")
        expected = expected_behavior(row)
        actual = actual_behavior(status)
        refused = status not in STATUS_ANSWERED
        accurate = bool(gold) and predicted == gold and not refused
        item = {
            "id": _cell(row, "id") or f"row_{index:03d}",
            "qa_type": _cell(row, "qa_type"),
            "source_type": _cell(row, "source_type"),
            "difficulty": _cell(row, "difficulty_cn") or _cell(row, "difficulty"),
            "question": _cell(row, "question"),
            "original_question": _cell(row, "question"),
            "submitted_question": query,
            "query_sent_to_system": query,
            "expected_behavior": expected,
            "actual_behavior": actual,
            "behavior_correct": expected == actual,
            "retrieval_coverage": retrieval_coverage(response),
            "options": {key: _cell(row, f"option_{key.lower()}") for key in OPTION_KEYS},
            "gold_answer": gold,
            "gold_answer_text": _cell(row, "answer_text"),
            "predicted_answer": predicted,
            "status": status,
            "refused": refused,
            "accurate": accurate,
            "refusal_reason": _refusal_reason(response),
            "source_title": _cell(row, "source_title"),
            "file_label": _cell(row, "file_label"),
            "gold_evidence": _cell(row, "evidence"),
            "option_features": _option_features(response),
            "system_answer": response.get("answer", ""),
            "system_error": error or response.get("error", ""),
            "elapsed_ms": elapsed_ms,
        }
        results.append(item)
        print(f"[{index}/{len(selected_rows)}] {item['id']} status={status} predicted={predicted or '-'} gold={gold or '-'} accurate={accurate}")
        if delay > 0:
            time.sleep(delay)
    return results


def build_summary(results: list[dict[str, Any]]) -> dict[str, Any]:
    total = len(results)
    refused = sum(bool(r["refused"]) for r in results)
    accurate = sum(bool(r["accurate"]) for r in results)
    answered = total - refused
    inaccurate = sum((not r["accurate"]) and not r["refused"] for r in results)
    by_reason: dict[str, int] = {}
    by_type: dict[str, dict[str, int]] = {}
    for item in results:
        reason = item["refusal_reason"] or "未说明"
        if item["refused"]:
            by_reason[reason] = by_reason.get(reason, 0) + 1
        qa_type = item["qa_type"] or "未分类"
        bucket = by_type.setdefault(qa_type, {"total": 0, "answered": 0, "refused": 0, "accurate": 0, "inaccurate": 0})
        bucket["total"] += 1
        bucket["answered" if not item["refused"] else "refused"] += 1
        bucket["accurate" if item["accurate"] else "inaccurate"] += 1
    return {
        "total": total,
        "answered": answered,
        "refused": refused,
        "accurate": accurate,
        "inaccurate_answered": inaccurate,
        "answer_rate": round(answered / total, 4) if total else 0.0,
        "accuracy_over_all": round(accurate / total, 4) if total else 0.0,
        "accuracy_over_answered": round(accurate / answered, 4) if answered else 0.0,
        "retrieval_metrics": build_retrieval_metrics(results),
        "behavior_metrics": build_behavior_metrics(results),
        "refusal_reasons": dict(sorted(by_reason.items(), key=lambda kv: (-kv[1], kv[0]))),
        "by_qa_type": by_type,
        "refused_ids": [r["id"] for r in results if r["refused"]],
        "inaccurate_ids": [r["id"] for r in results if (not r["accurate"]) and not r["refused"]],
    }


def write_reports(results: list[dict[str, Any]], summary: dict[str, Any], output_dir: Path) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    with (output_dir / "qa_results.jsonl").open("w", encoding="utf-8") as handle:
        for item in results:
            handle.write(json.dumps(item, ensure_ascii=False) + "\n")
    with (output_dir / "qa_summary.json").open("w", encoding="utf-8") as handle:
        json.dump(summary, handle, ensure_ascii=False, indent=2)
    columns = [
        "id", "qa_type", "source_type", "original_question", "submitted_question",
        "expected_behavior", "actual_behavior", "behavior_correct", "gold_answer", "predicted_answer",
        "status", "refused", "accurate", "refusal_reason", "source_title", "file_label",
        "system_answer", "system_error", "elapsed_ms",
    ]
    with (output_dir / "qa_results.csv").open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(results)


def main() -> None:
    parser = argparse.ArgumentParser(description="用 QA数据.xlsx 评测项目 RAG 问答系统")
    parser.add_argument("--qa", type=Path, default=_default_qa_path(), help="QA xlsx 路径")
    parser.add_argument("--sheet", default=None, help="工作表名称，默认第一个工作表")
    parser.add_argument("--output-dir", type=Path, default=PROJECT_ROOT / "reports" / "qa_eval")
    parser.add_argument("--limit", type=int, default=None, help="只测试前 N 条，便于先试跑")
    parser.add_argument("--delay", type=float, default=0.0, help="每题之间等待秒数")
    parser.add_argument("--local", action="store_true", help="禁用 DeepSeek，使用系统本地兜底生成器")
    args = parser.parse_args()
    if not args.qa.exists():
        parser.error(f"QA 文件不存在：{args.qa}")
    rows = load_rows(args.qa, args.sheet)
    results = evaluate(rows, local=args.local, limit=args.limit, delay=args.delay)
    summary = build_summary(results)
    write_reports(results, summary, args.output_dir)
    print("\n=== 评测汇总 ===")
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    print(f"报告目录：{args.output_dir.resolve()}")


if __name__ == "__main__":
    main()
